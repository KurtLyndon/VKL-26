import csv
import re
import unicodedata
from io import BytesIO, StringIO

from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.models import (
    Target,
    TargetAttributeDefinition,
    TargetAttributeValue,
    TargetGroup,
    TargetGroupMapping,
)
from app.schemas.resources import (
    TargetAttributeAssignmentUpdateRequest,
    TargetDetailRead,
    TargetGroupAssignmentUpdateRequest,
    TargetImportResponse,
)

BASE_TARGET_COLUMNS = {
    "code": "code",
    "ma": "code",
    "name": "name",
    "ten": "name",
    "target_name": "name",
    "ip": "ip_range",
    "ips": "ip_range",
    "ip_range": "ip_range",
    "iprange": "ip_range",
    "ip_range_cidr": "ip_range",
    "cidr": "ip_range",
    "range": "ip_range",
    "dai_ip": "ip_range",
    "ip_dai": "ip_range",
    "domain": "domain",
    "target_type": "target_type",
    "type": "target_type",
    "description": "description",
    "mo_ta": "description",
}


def normalize_key(value: str) -> str:
    ascii_text = (
        unicodedata.normalize("NFKD", value or "").encode("ascii", "ignore").decode("ascii").strip().lower()
    )
    return re.sub(r"[^a-z0-9]+", "_", ascii_text).strip("_")


def make_code(prefix: str, value: str, fallback_index: int) -> str:
    normalized = normalize_key(value)
    if not normalized:
        normalized = f"{prefix}_{fallback_index}"
    return f"{prefix}_{normalized}"[:50]


def detect_ip_entry_type(ip_range: str | None) -> str:
    if not ip_range:
        return "empty"
    if "/" in ip_range:
        return "cidr"
    if "-" in ip_range:
        return "range"
    if "," in ip_range or "\n" in ip_range:
        return "list"
    return "single"


def list_targets_enriched(db: Session) -> list[TargetDetailRead]:
    targets = db.scalars(select(Target).order_by(Target.id.desc())).all()
    definitions = db.scalars(select(TargetAttributeDefinition).order_by(TargetAttributeDefinition.attribute_name.asc())).all()
    definition_map = {item.id: item for item in definitions}

    values = db.scalars(select(TargetAttributeValue)).all()
    value_map: dict[int, dict[int, str | None]] = {}
    for row in values:
        value_map.setdefault(row.target_id, {})[row.attribute_definition_id] = row.value_text

    groups = db.scalars(select(TargetGroup).order_by(TargetGroup.name.asc())).all()
    group_map = {item.id: item for item in groups}
    mappings = db.scalars(select(TargetGroupMapping)).all()
    target_group_ids: dict[int, list[int]] = {}
    for row in mappings:
        target_group_ids.setdefault(row.target_id, []).append(row.target_group_id)

    items: list[TargetDetailRead] = []
    for target in targets:
        group_ids = target_group_ids.get(target.id, [])
        items.append(
            TargetDetailRead(
                id=target.id,
                code=target.code,
                name=target.name,
                target_type=target.target_type,
                ip_range=target.ip_range,
                domain=target.domain,
                description=target.description,
                created_at=target.created_at,
                updated_at=target.updated_at,
                ip_entry_type=detect_ip_entry_type(target.ip_range),
                attribute_values=[
                    {
                        "attribute_definition_id": definition.id,
                        "attribute_code": definition.attribute_code,
                        "attribute_name": definition.attribute_name,
                        "data_type": definition.data_type,
                        "value_text": value_map.get(target.id, {}).get(definition.id),
                    }
                    for definition in definitions
                ],
                group_ids=group_ids,
                groups=[
                    {
                        "id": group.id,
                        "code": group.code,
                        "name": group.name,
                    }
                    for group_id in group_ids
                    if (group := group_map.get(group_id))
                ],
            )
        )
    return items


def create_target(db: Session, payload: dict) -> Target:
    code = payload.get("code") or make_code("target", payload.get("name", ""), fallback_index=1)
    counter = 1
    base_code = code
    while db.scalar(select(Target).where(Target.code == code)):
        counter += 1
        code = f"{base_code[:45]}_{counter}"[:50]

    item = Target(
        code=code,
        name=payload["name"],
        target_type=payload.get("target_type") or "network",
        ip_range=payload.get("ip_range"),
        domain=payload.get("domain"),
        description=payload.get("description"),
    )
    db.add(item)
    db.commit()
    db.refresh(item)
    return item


def update_target(db: Session, target: Target, payload: dict) -> Target:
    for field in ("name", "target_type", "ip_range", "domain", "description"):
        if field in payload:
            setattr(target, field, payload[field])
    if payload.get("code") and payload["code"] != target.code:
        target.code = payload["code"]
    db.commit()
    db.refresh(target)
    return target


def delete_target(db: Session, target: Target) -> None:
    for row in db.scalars(select(TargetAttributeValue).where(TargetAttributeValue.target_id == target.id)).all():
        db.delete(row)
    for row in db.scalars(select(TargetGroupMapping).where(TargetGroupMapping.target_id == target.id)).all():
        db.delete(row)
    db.delete(target)
    db.commit()


def delete_target_attribute_definition(db: Session, definition: TargetAttributeDefinition) -> None:
    for row in db.scalars(
        select(TargetAttributeValue).where(TargetAttributeValue.attribute_definition_id == definition.id)
    ).all():
        db.delete(row)
    db.delete(definition)
    db.commit()


def delete_target_group(db: Session, group: TargetGroup) -> None:
    for row in db.scalars(select(TargetGroupMapping).where(TargetGroupMapping.target_group_id == group.id)).all():
        db.delete(row)
    db.delete(group)
    db.commit()


def update_target_attribute_assignments(
    db: Session, target_id: int, payload: TargetAttributeAssignmentUpdateRequest
) -> list[TargetAttributeValue]:
    existing = db.scalars(select(TargetAttributeValue).where(TargetAttributeValue.target_id == target_id)).all()
    existing_map = {item.attribute_definition_id: item for item in existing}
    for item in payload.items:
        row = existing_map.get(item.attribute_definition_id)
        if row:
            row.value_text = item.value_text
        else:
            db.add(
                TargetAttributeValue(
                    target_id=target_id,
                    attribute_definition_id=item.attribute_definition_id,
                    value_text=item.value_text,
                )
            )
    db.commit()
    return db.scalars(select(TargetAttributeValue).where(TargetAttributeValue.target_id == target_id)).all()


def update_target_group_assignments(
    db: Session, target_id: int, payload: TargetGroupAssignmentUpdateRequest
) -> list[TargetGroupMapping]:
    existing = db.scalars(select(TargetGroupMapping).where(TargetGroupMapping.target_id == target_id)).all()
    existing_ids = {item.target_group_id for item in existing}
    desired_ids = set(payload.target_group_ids)

    for item in existing:
        if item.target_group_id not in desired_ids:
            db.delete(item)

    for group_id in desired_ids - existing_ids:
        db.add(TargetGroupMapping(target_id=target_id, target_group_id=group_id))

    db.commit()
    return db.scalars(select(TargetGroupMapping).where(TargetGroupMapping.target_id == target_id)).all()


def parse_target_import_file(file_name: str, content: bytes) -> list[dict[str, str | None]]:
    suffix = file_name.lower().rsplit(".", 1)[-1] if "." in file_name else ""
    if suffix == "csv":
        text = content.decode("utf-8-sig")
        return list(csv.DictReader(StringIO(text)))

    if suffix in {"xlsx", "xlsm"}:
        workbook = load_workbook(filename=BytesIO(content), read_only=True, data_only=True)
        worksheet = workbook[workbook.sheetnames[0]]
        rows = list(worksheet.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
        records: list[dict[str, str | None]] = []
        for row in rows[1:]:
            if not any(cell not in (None, "") for cell in row):
                continue
            records.append(
                {
                    headers[index]: None if index >= len(row) or row[index] is None else str(row[index]).strip()
                    for index in range(len(headers))
                    if headers[index]
                }
            )
        return records

    raise ValueError("Chỉ hỗ trợ file CSV hoặc XLSX.")


def import_targets_from_file(db: Session, file_name: str, content: bytes) -> TargetImportResponse:
    rows = parse_target_import_file(file_name, content)
    definitions = db.scalars(select(TargetAttributeDefinition)).all()
    definition_by_code = {item.attribute_code: item for item in definitions}
    created_attribute_definitions = 0
    created_targets = 0
    updated_targets = 0

    for index, row in enumerate(rows, start=1):
        normalized = {normalize_key(key): value for key, value in row.items()}
        base_payload: dict[str, str | None] = {}
        extra_attributes: dict[str, str | None] = {}

        for key, value in normalized.items():
            mapped = BASE_TARGET_COLUMNS.get(key)
            if mapped:
                base_payload[mapped] = value
            else:
                extra_attributes[key] = value

        name = base_payload.get("name")
        if not name:
            continue

        code = base_payload.get("code") or make_code("target", name, fallback_index=index)
        target = db.scalar(select(Target).where(Target.code == code))
        if target is None:
            target = Target(
                code=code,
                name=name,
                target_type=base_payload.get("target_type") or "network",
                ip_range=base_payload.get("ip_range"),
                domain=base_payload.get("domain"),
                description=base_payload.get("description"),
            )
            db.add(target)
            db.flush()
            created_targets += 1
        else:
            target.name = name
            target.target_type = base_payload.get("target_type") or target.target_type
            target.ip_range = base_payload.get("ip_range")
            target.domain = base_payload.get("domain")
            target.description = base_payload.get("description")
            updated_targets += 1

        for attribute_code, attribute_value in extra_attributes.items():
            definition = definition_by_code.get(attribute_code)
            if definition is None:
                definition = TargetAttributeDefinition(
                    attribute_code=attribute_code[:50],
                    attribute_name=(attribute_code.replace("_", " ").title())[:255],
                    data_type="text",
                    is_required=False,
                    default_value=None,
                    description="Tự tạo từ file import target.",
                )
                db.add(definition)
                db.flush()
                definition_by_code[attribute_code] = definition
                created_attribute_definitions += 1

            value_row = db.scalar(
                select(TargetAttributeValue).where(
                    TargetAttributeValue.target_id == target.id,
                    TargetAttributeValue.attribute_definition_id == definition.id,
                )
            )
            if value_row is None:
                db.add(
                    TargetAttributeValue(
                        target_id=target.id,
                        attribute_definition_id=definition.id,
                        value_text=attribute_value,
                    )
                )
            else:
                value_row.value_text = attribute_value

    db.commit()
    return TargetImportResponse(
        imported_targets=created_targets + updated_targets,
        created_targets=created_targets,
        updated_targets=updated_targets,
        created_attribute_definitions=created_attribute_definitions,
        source_file_name=file_name,
    )
