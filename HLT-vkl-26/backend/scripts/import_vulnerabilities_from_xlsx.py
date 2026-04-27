from __future__ import annotations

import argparse
from itertools import chain
import re
import unicodedata
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import select

from app.db.session import SessionLocal
from app.models import Vulnerability, VulnerabilityScript
from app.services.poc_repository import store_poc_copy
from database.migrations import apply_migrations


SCRIPT_EXTENSIONS = {".py", ".sh", ".ps1", ".ps", ".bat", ".cmd", ".exe"}


def normalize_header(value: str | None) -> str:
    text = (value or "").strip().lower()
    text = text.replace("đ", "d").replace("Đ", "D")
    text = unicodedata.normalize("NFKD", text)
    text = "".join(char for char in text if not unicodedata.combining(char))
    return "".join(char for char in text if char.isalnum())


def sanitize_text(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).replace("\r\n", "\n").replace("\r", "\n").strip()
    return text or None


def parse_level(value: object) -> int:
    if value is None:
        return 0
    if isinstance(value, (int, float)):
        return int(value)

    text = sanitize_text(value) or ""
    lowered = normalize_header(text)
    mapping = {
        "thong tin": 0,
        "thap": 1,
        "trung binh": 2,
        "cao": 3,
        "nghiem trong": 4,
    }
    if lowered in mapping:
        return mapping[lowered]

    match = re.search(r"\d+", text)
    return int(match.group()) if match else 0


def classify_poc(value: str | None) -> tuple[str | None, str | None]:
    if not value:
        return None, None
    candidate = value.strip()
    suffix = Path(candidate).suffix.lower()
    if suffix in SCRIPT_EXTENSIONS and "\n" not in candidate:
        return candidate, None
    return None, candidate


def read_script_content(file_path: Path) -> str:
    return file_path.read_text(encoding="utf-8", errors="ignore")


def upsert_vulnerability_script(db, vulnerability_id: int, poc_source_file: Path) -> None:
    script_name = poc_source_file.name
    script_type = poc_source_file.suffix.lower().lstrip(".") or "txt"
    script_content = read_script_content(poc_source_file)

    existing = db.scalar(
        select(VulnerabilityScript).where(VulnerabilityScript.vulnerability_id == vulnerability_id)
    )
    if existing is None:
        db.add(
            VulnerabilityScript(
                vulnerability_id=vulnerability_id,
                script_name=script_name,
                script_type=script_type,
                script_content=script_content,
                version="imported",
                is_active=True,
            )
        )
        return

    existing.script_name = script_name
    existing.script_type = script_type
    existing.script_content = script_content
    existing.version = "imported"
    existing.is_active = True


def derive_title(code: str, threat: str | None) -> str:
    if not threat:
        return code

    text = threat.strip()
    text = re.sub(r"^[\-\u2022\s]+", "", text)
    if text.lower().startswith(code.lower()):
        text = text[len(code) :].lstrip(" :-")
    first_sentence = re.split(r"(?<=[.!?])\s+", text, maxsplit=1)[0].strip()
    title = first_sentence[:255].strip()
    return title or code


HEADER_ALIASES = {
    "code": ("ma", "macve", "cve", "vulnerabilitycode"),
    "level": ("mucdo", "mucdoruiro", "severity", "risklevel", "level"),
    "threat": ("nguycomatattt", "nguyco", "mota", "threat", "description"),
    "proposal": ("kiennghidexuat", "kiennghi", "dexuat", "recommendation", "proposal"),
    "poc": ("kiemchung", "poc", "proof", "proofconcept"),
}


def locate_columns(header_row: tuple[object, ...]) -> dict[str, int]:
    normalized = {
        normalize_header(str(value) if value is not None else ""): index
        for index, value in enumerate(header_row)
    }
    columns: dict[str, int] = {}

    for target, aliases in HEADER_ALIASES.items():
        for alias in aliases:
            if alias in normalized:
                columns[target] = normalized[alias]
                break

    required_labels = {
        "code": "ma",
        "level": "muc do",
        "threat": "nguy co mat attt",
        "proposal": "kien nghi, de xuat",
    }
    for target, source in required_labels.items():
        if target not in columns:
            raise ValueError(f"Khong tim thay cot bat buoc: {source}")

    return columns


def locate_header_row(rows: list[tuple[object, ...]]) -> tuple[tuple[object, ...], dict[str, int], int]:
    last_error: ValueError | None = None
    for index, row in enumerate(rows):
        if not any(cell is not None and str(cell).strip() for cell in row):
            continue
        try:
            columns = locate_columns(row)
            return row, columns, index
        except ValueError as exc:
            last_error = exc
            continue

    if last_error is not None:
        raise last_error
    raise ValueError("Khong tim thay dong header hop le trong workbook")


def import_workbook(workbook_path: Path) -> tuple[int, int, int]:
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    preview_rows: list[tuple[object, ...]] = []
    for _ in range(10):
        try:
            preview_rows.append(next(rows))
        except StopIteration:
            break
    _, columns, header_index = locate_header_row(preview_rows)
    data_rows = iter(preview_rows[header_index + 1 :])

    apply_migrations()

    created = 0
    updated = 0
    copied_poc_files = 0

    db = SessionLocal()
    try:
        for row in chain(data_rows, rows):
            code = sanitize_text(row[columns["code"]])
            if not code:
                continue

            threat = sanitize_text(row[columns["threat"]])
            proposal = sanitize_text(row[columns["proposal"]])
            level = parse_level(row[columns["level"]])
            poc_raw = sanitize_text(row[columns["poc"]]) if "poc" in columns else None
            poc_file_name, poc_text = classify_poc(poc_raw)
            poc_source_file = None
            if poc_file_name:
                candidate = workbook_path.parent / poc_file_name
                if candidate.exists():
                    poc_source_file = candidate
                else:
                    poc_text = poc_raw
                    poc_file_name = None

            vulnerability = db.scalar(select(Vulnerability).where(Vulnerability.code == code))
            payload = {
                "title": derive_title(code, threat),
                "level": level,
                "threat": threat,
                "proposal": proposal,
                "poc_file_name": poc_file_name,
                "poc_text": poc_text,
                "description": threat,
            }

            if vulnerability is None:
                vulnerability = Vulnerability(code=code, **payload)
                db.add(vulnerability)
                created += 1
            else:
                for field, value in payload.items():
                    setattr(vulnerability, field, value)
                updated += 1

            db.flush()

            if poc_source_file is not None:
                copied_path = store_poc_copy(poc_source_file, code)
                vulnerability.poc_file_name = copied_path.name
                upsert_vulnerability_script(db, vulnerability.id, poc_source_file)
                copied_poc_files += 1

        db.commit()
    finally:
        db.close()

    return created, updated, copied_poc_files


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Import vulnerability/CVE data from an XLSX workbook into HLT.")
    parser.add_argument("workbook_path", help="Path to the XLSX workbook.")
    return parser


def main() -> int:
    parser = build_parser()
    args = parser.parse_args()
    workbook_path = Path(args.workbook_path)
    if not workbook_path.exists():
        raise SystemExit(f"Workbook not found: {workbook_path}")

    created, updated, copied_poc_files = import_workbook(workbook_path)
    print(
        f"Imported vulnerabilities from {workbook_path.name}: "
        f"created={created}, updated={updated}, poc_files_copied={copied_poc_files}"
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
