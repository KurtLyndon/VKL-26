from __future__ import annotations

import argparse
import unicodedata
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy import select

from app.db.session import SessionLocal
from app.models import Target, TargetAttributeDefinition, TargetAttributeValue, TargetGroup, TargetGroupMapping
from app.services.targets import make_code, normalize_target_ip_range
from database.migrations import apply_migrations


HEADER_ALIASES = {
    "name": ("donvi",),
    "ip_range": ("daiip",),
    "dv_cap_1": ("dvcap1",),
    "cap": ("cap",),
    "lv1": ("lv1", "cocap1trendiaban"),
    "non_lv1_lv2": ("nonlv1lv2", "cap2khongcocap1trendiaban"),
}


ATTRIBUTE_DEFINITIONS = {
    "dv_cap_1": ("dv_cap_1", "ĐV Cấp 1", "text"),
    "cap": ("cap", "Cấp", "text"),
    "lv1": ("lv1", "Lv1", "boolean"),
    "non_lv1_lv2": ("non_lv1_lv2", "non-lv1 Lv2", "boolean"),
}
TARGET_GROUP_CODE = "TRTA2"
TARGET_GROUP_NAME = "TRTA2"


def normalize_header(value: object) -> str:
    text = str(value or "").strip().lower().replace("đ", "d").replace("Đ", "D")
    text = unicodedata.normalize("NFKD", text)
    text = "".join(char for char in text if not unicodedata.combining(char))
    return "".join(char for char in text if char.isalnum())


def sanitize_text(value: object) -> str | None:
    if value is None:
        return None
    text = str(value).replace("\r\n", "\n").replace("\r", "\n").strip()
    return text or None


def to_boolean_text(value: object) -> str | None:
    if value is None:
        return None
    if isinstance(value, bool):
        return "true" if value else "false"

    text = sanitize_text(value)
    if not text:
        return None

    lowered = normalize_header(text)
    if lowered in {"true", "yes", "co", "x", "1"}:
        return "true"
    if lowered in {"false", "no", "khong", "0"}:
        return "false"
    return None


def locate_columns(header_row: tuple[object, ...]) -> dict[str, int]:
    normalized = {
        normalize_header(value): index
        for index, value in enumerate(header_row)
    }
    columns: dict[str, int] = {}
    for target_key, aliases in HEADER_ALIASES.items():
        for alias in aliases:
            if alias in normalized:
                columns[target_key] = normalized[alias]
                break

    required = {"name": "Đơn vị", "ip_range": "Dải IP"}
    for key, label in required.items():
        if key not in columns:
            raise ValueError(f"Không tìm thấy cột bắt buộc: {label}")
    return columns


def ensure_attribute_definition(db, code: str, name: str, data_type: str) -> TargetAttributeDefinition:
    definition = db.scalar(
        select(TargetAttributeDefinition).where(TargetAttributeDefinition.attribute_code == code)
    )
    if definition is None:
        definition = TargetAttributeDefinition(
            attribute_code=code,
            attribute_name=name,
            data_type=data_type,
            is_required=False,
            default_value=None,
            description="Sinh từ seed/import target basing workbook.",
        )
        db.add(definition)
        db.flush()
    else:
        definition.attribute_name = name
        definition.data_type = data_type
    return definition


def ensure_target_group(db) -> TargetGroup:
    group = db.scalar(select(TargetGroup).where(TargetGroup.code == TARGET_GROUP_CODE))
    if group is None:
        group = TargetGroup(
            code=TARGET_GROUP_CODE,
            name=TARGET_GROUP_NAME,
            description="Nhóm target import từ workbook basing STMNC.",
        )
        db.add(group)
        db.flush()
    else:
        group.name = TARGET_GROUP_NAME
    return group


def ensure_target_group_mapping(db, target_id: int, group_id: int) -> None:
    existing = db.scalar(
        select(TargetGroupMapping).where(
            TargetGroupMapping.target_id == target_id,
            TargetGroupMapping.target_group_id == group_id,
        )
    )
    if existing is None:
        db.add(TargetGroupMapping(target_id=target_id, target_group_id=group_id))


def upsert_attribute_value(db, target_id: int, definition_id: int, value_text: str | None) -> None:
    existing = db.scalar(
        select(TargetAttributeValue).where(
            TargetAttributeValue.target_id == target_id,
            TargetAttributeValue.attribute_definition_id == definition_id,
        )
    )
    if existing is None:
        db.add(
            TargetAttributeValue(
                target_id=target_id,
                attribute_definition_id=definition_id,
                value_text=value_text,
            )
        )
        return
    existing.value_text = value_text


def import_workbook(workbook_path: Path) -> tuple[int, int]:
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return 0, 0

    columns = locate_columns(rows[0])
    apply_migrations()

    created = 0
    updated = 0
    db = SessionLocal()
    try:
        definitions = {
            key: ensure_attribute_definition(db, *config)
            for key, config in ATTRIBUTE_DEFINITIONS.items()
        }
        target_group = ensure_target_group(db)

        for index, row in enumerate(rows[1:], start=1):
            name = sanitize_text(row[columns["name"]])
            if not name:
                continue

            code = make_code("target", name, fallback_index=index)
            target = db.scalar(select(Target).where(Target.code == code))
            normalized_ip_range = normalize_target_ip_range(sanitize_text(row[columns["ip_range"]]))

            if target is None:
                target = Target(
                    code=code,
                    name=name,
                    target_type="network",
                    ip_range=normalized_ip_range,
                    domain=None,
                    description="mô tả mục tiêu",
                )
                db.add(target)
                db.flush()
                created += 1
            else:
                target.name = name
                target.target_type = "network"
                target.ip_range = normalized_ip_range
                target.domain = None
                target.description = "mô tả mục tiêu"
                updated += 1

            ensure_target_group_mapping(db, target.id, target_group.id)

            upsert_attribute_value(
                db,
                target.id,
                definitions["dv_cap_1"].id,
                sanitize_text(row[columns["dv_cap_1"]]) if "dv_cap_1" in columns else None,
            )
            upsert_attribute_value(
                db,
                target.id,
                definitions["cap"].id,
                sanitize_text(row[columns["cap"]]) if "cap" in columns else None,
            )
            upsert_attribute_value(
                db,
                target.id,
                definitions["lv1"].id,
                to_boolean_text(row[columns["lv1"]]) if "lv1" in columns else None,
            )
            upsert_attribute_value(
                db,
                target.id,
                definitions["non_lv1_lv2"].id,
                to_boolean_text(row[columns["non_lv1_lv2"]]) if "non_lv1_lv2" in columns else None,
            )

        db.commit()
    finally:
        db.close()

    return created, updated


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Import target basing workbook into HLT.")
    parser.add_argument("workbook_path", help="Path to the XLSX workbook.")
    return parser


def main() -> int:
    parser = build_parser()
    args = parser.parse_args()
    workbook_path = Path(args.workbook_path)
    if not workbook_path.exists():
        raise SystemExit(f"Workbook not found: {workbook_path}")

    created, updated = import_workbook(workbook_path)
    print(f"Imported targets from {workbook_path.name}: created={created}, updated={updated}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
